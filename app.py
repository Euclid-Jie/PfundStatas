from __future__ import annotations

from io import BytesIO
import sqlite3
from pathlib import Path

from flask import Flask, jsonify, render_template, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from config import SQLITE_PATH, WEB_PORT
from update_data import START_DATE, init_db


ROOT = Path(__file__).resolve().parent
app = Flask(__name__, template_folder=str(ROOT), static_folder=str(ROOT), static_url_path="")

with sqlite3.connect(SQLITE_PATH) as _conn:
    init_db(_conn)


def get_conn() -> sqlite3.Connection:
    conn = sqlite3.connect(SQLITE_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def rows_to_dicts(rows):
    return [dict(row) for row in rows]


def sum_months(month_rows, months):
    month_set = set(months)
    return sum(int(row["record_count"] or 0) for row in month_rows if row["month"] in month_set)


def load_manager_rows(conn: sqlite3.Connection, manager_keyword: str = ""):
    where = ["put_on_record_date >= ?"]
    params = [START_DATE]
    if manager_keyword:
        where.append("(manager_short_name LIKE ? OR manager_name LIKE ?)")
        like = f"%{manager_keyword}%"
        params.extend([like, like])

    return conn.execute(
        f"""
        SELECT
            substr(put_on_record_date, 1, 7) AS month,
            COALESCE(NULLIF(manager_short_name, ''), manager_name) AS manager_name,
            COUNT(*) AS record_count
        FROM records
        WHERE {' AND '.join(where)}
        GROUP BY month, manager_name
        ORDER BY month ASC, manager_name ASC
        """,
        params,
    ).fetchall()


def build_manager_pivot(rows, start_year: int = 2024):
    months_set = set()
    manager_map = {}
    summary_months = {}

    for row in rows:
        month = row["month"]
        if not month or int(month[:4]) < start_year:
            continue
        months_set.add(month)

        manager = row["manager_name"] or "-"
        count = int(row["record_count"] or 0)
        manager_map.setdefault(manager, {"manager_name": manager, "months": {}, "all_total": 0})
        manager_map[manager]["months"][month] = count
        manager_map[manager]["all_total"] += count
        summary_months[month] = summary_months.get(month, 0) + count

    months = sorted(months_set)
    years = sorted({m[:4] for m in months})
    recent_months = months[-12:]
    latest_year = years[-1] if years else None
    latest_year_str = str(latest_year) if latest_year else ""

    summary_row = {
        "manager_name": "当月求和",
        "months": summary_months,
        "ytd": sum(v for m, v in summary_months.items() if m.startswith(latest_year_str)),
        "total": sum(summary_months.get(m, 0) for m in recent_months),
    }

    pivot_rows = []
    for manager in manager_map.values():
        mm = manager["months"]
        recent_total = sum(mm.get(m, 0) for m in recent_months)
        pivot_rows.append({
            "manager_name": manager["manager_name"],
            "months": {m: mm.get(m, 0) for m in months},
            "ytd": sum(v for m, v in mm.items() if m.startswith(latest_year_str)),
            "total": recent_total,
            "all_total": manager["all_total"],
        })

    pivot_rows.sort(key=lambda item: (-item["total"], item["manager_name"]))
    return {"years": years, "months": months, "summary_row": summary_row, "rows": pivot_rows}


def build_manager_pivot_workbook(pivot: dict) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "月度求和"

    months = pivot.get("months", [])
    summary_row = pivot.get("summary_row", {})
    rows = pivot.get("rows", [])

    sheet["A1"] = "管理人月度求和（2024 年起）"
    sheet["A2"] = "仅统计 2024 年起数据；最后一列为近一年合计"
    sheet["A1"].font = Font(bold=True, size=14)
    sheet["A2"].font = Font(color="666666", italic=True)

    headers = ["管理人", *months, "YTD", "近一年"]
    header_row = 4
    for column, value in enumerate(headers, start=1):
        cell = sheet.cell(row=header_row, column=column, value=value)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F2937")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    data_rows = [
        {
            "manager_name": summary_row.get("manager_name", "当月求和"),
            "months": summary_row.get("months", {}),
            "ytd": summary_row.get("ytd", 0),
            "total": summary_row.get("total", 0),
        }
    ]
    data_rows.extend(rows)

    for row_index, item in enumerate(data_rows, start=header_row + 1):
        is_summary = row_index == header_row + 1
        name_cell = sheet.cell(row=row_index, column=1, value=item.get("manager_name", "-"))
        name_cell.font = Font(bold=is_summary)
        name_cell.alignment = Alignment(horizontal="left")
        for offset, month in enumerate(months, start=2):
            cell = sheet.cell(row=row_index, column=offset, value=int(item.get("months", {}).get(month, 0) or 0))
            cell.alignment = Alignment(horizontal="center")
        ytd_cell = sheet.cell(row=row_index, column=2 + len(months), value=int(item.get("ytd", 0) or 0))
        total_cell = sheet.cell(row=row_index, column=3 + len(months), value=int(item.get("total", 0) or 0))
        ytd_cell.font = Font(bold=is_summary)
        total_cell.font = Font(bold=is_summary)
        ytd_cell.alignment = Alignment(horizontal="center")
        total_cell.alignment = Alignment(horizontal="center")

    sheet.freeze_panes = "B5"
    sheet.auto_filter.ref = f"A4:{sheet.cell(row=4, column=3 + len(months)).column_letter}4"
    sheet.column_dimensions["A"].width = 24
    for idx in range(2, 2 + len(months)):
        sheet.column_dimensions[sheet.cell(row=header_row, column=idx).column_letter].width = 10
    sheet.column_dimensions[sheet.cell(row=header_row, column=2 + len(months)).column_letter].width = 12
    sheet.column_dimensions[sheet.cell(row=header_row, column=3 + len(months)).column_letter].width = 12

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/dashboard")
def dashboard():
    manager_keyword = request.args.get("manager_q", "").strip()

    with get_conn() as conn:
        summary = conn.execute(
            """
            SELECT
                COUNT(*) AS total_records,
                COUNT(DISTINCT manager_name) AS unique_managers,
                MAX(put_on_record_date) AS latest_record_date
            FROM records
            """
        ).fetchone()
        weekly = conn.execute(
            """
            SELECT
                date(put_on_record_date, '+' || ((5 - strftime('%w', put_on_record_date) + 7) % 7) || ' days') AS week,
                COUNT(*) AS record_count
            FROM records
            GROUP BY week
            ORDER BY week
            """
        ).fetchall()
        monthly = conn.execute(
            """
            SELECT
                substr(put_on_record_date, 1, 7) AS month,
                COUNT(*) AS record_count
            FROM records
            GROUP BY month
            ORDER BY month
            """
        ).fetchall()
        monthly_rows = rows_to_dicts(monthly)
        month_keys = [row["month"] for row in monthly_rows]
        recent_months = month_keys[-12:]
        current_year = month_keys[-1][:4] if month_keys else None
        ytd_months = [month for month in recent_months if current_year and month.startswith(current_year)]
        manager_rows = load_manager_rows(conn, manager_keyword)
        manager_pivot = build_manager_pivot(manager_rows, 2024)

    summary_payload = dict(summary)
    summary_payload.update(
        recent_year_records=sum_months(monthly_rows, recent_months),
        ytd_records=sum_months(monthly_rows, ytd_months),
        latest_week=weekly[-1]["week"] if weekly else None,
        latest_week_records=weekly[-1]["record_count"] if weekly else 0,
        latest_month=monthly[-1]["month"] if monthly else None,
        latest_month_records=monthly[-1]["record_count"] if monthly else 0,
    )

    return jsonify(
        summary=summary_payload,
        weekly_series=rows_to_dicts(weekly),
        monthly_series=monthly_rows,
        manager_pivot=manager_pivot,
    )


@app.route("/api/manager-pivot.xlsx")
def manager_pivot_xlsx():
    manager_keyword = request.args.get("manager_q", "").strip()

    with get_conn() as conn:
        manager_rows = load_manager_rows(conn, manager_keyword)
        pivot = build_manager_pivot(manager_rows, 2024)

    buffer = build_manager_pivot_workbook(pivot)
    return send_file(
        buffer,
        as_attachment=True,
        download_name="manager-monthly-pivot.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/api/records")
def records():
    keyword = request.args.get("q", "").strip()
    page = max(int(request.args.get("page", 1)), 1)
    size = min(max(int(request.args.get("size", 12)), 1), 100)

    conditions = []
    params = []
    if keyword:
        conditions.append("(fund_name LIKE ? OR manager_name LIKE ?)")
        like = f"%{keyword}%"
        params.extend([like, like])

    where = f"WHERE {' AND '.join(conditions)}" if conditions else ""
    offset = (page - 1) * size
    with get_conn() as conn:
        total = conn.execute(
            f"SELECT COUNT(*) AS count FROM records {where}",
            params,
        ).fetchone()["count"]
        rows = conn.execute(
            f"""
            SELECT
                fund_no AS fundNo,
                fund_name AS fundName,
                manager_name AS managerName,
                COALESCE(manager_short_name, manager_name) AS managerShortName,
                manager_type AS managerType,
                working_state AS workingState,
                put_on_record_date AS putOnRecordDate,
                mandator_name AS mandatorName,
                register_no AS registerNo
            FROM records
            {where}
            ORDER BY put_on_record_date DESC, fund_no DESC
            LIMIT ? OFFSET ?
            """,
            [*params, size, offset],
        ).fetchall()
    return jsonify(total=total, page=page, size=size, items=rows_to_dicts(rows))


if __name__ == "__main__":
    app.run(host="127.0.0.1", port=WEB_PORT, debug=False)
